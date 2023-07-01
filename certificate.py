import cv2 as cv
import openpyxl
import img2pdf
from google.colab import files
import matplotlib.font_manager as fm

template_path = '/content/rotary-certificate_image.png'
details_path = '/content/names.csv.xlsx'

font_size = 2
font_color = (0, 0, 0)

coordinate_y_adjustment = 15
coordinate_x_adjustment = 7
coordinate_y_adjustment_below = -110
coordinate_x_adjustment_below = 120

obj = openpyxl.load_workbook(details_path)
sheet = obj.active
# Calculate the number of columns dynamically
num_rows = sheet.max_row
for i in range(2, num_rows + 1):
    get_name = sheet.cell(row=i, column=1)
    certi_name = get_name.value
    get_date = sheet.cell(row=i, column=2)
    date_id = get_date.value
    img = cv.imread(template_path)

    font =  cv.FONT_HERSHEY_COMPLEX_SMALL

    text_size = cv.getTextSize(certi_name, font, font_size, 2)[0]
    text_size_below = cv.getTextSize(date_id, font, font_size, 2)[0]

    text_x = (img.shape[1] - text_size[0]) // 2 + coordinate_x_adjustment
    text_y = (img.shape[0] + text_size[1]) // 2 - coordinate_y_adjustment
    text_x = int(text_x)
    text_y = int(text_y)
    cv.putText(img, certi_name, (text_x, text_y), font, font_size, font_color, 2)

    text_x_below = (img.shape[1] - text_size_below[0]) // 2 + coordinate_x_adjustment_below
    text_y_below = (img.shape[0] + text_size_below[1]) // 2 - coordinate_y_adjustment_below
    text_x_below = int(text_x_below)
    text_y_below = int(text_y_below)
    cv.putText(img, date_id, (text_x_below, text_y_below), font, font_size, font_color, 2)

    # Create PDF page from image
    _, temp_img = cv.imencode('.png', img)
    pdf_page = img2pdf.convert(temp_img.tobytes())

    # Save PDF page to a file
    output_pdf_path = f'/content/certificates/{certi_name}-CERT-.pdf'
    with open(output_pdf_path, 'wb') as f:
        f.write(pdf_page)

    # Download the PDF file
    files.download(output_pdf_path)
